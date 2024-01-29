import openpyxl

def print_sheet_contents(sheet):
    # 각 셀의 내용 출력
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        row_values = [cell.value for cell in row if cell.value is not None]
        print(row_values)

def read_and_write_excel(input_excel_path, output_excel_path, column_index):
    # 기존 엑셀 파일 열기
    input_wb = openpyxl.load_workbook(input_excel_path)
    input_sheet = input_wb.active

    # 엑셀 파일의 내용 출력
    print_sheet_contents(input_sheet)

    # for row in input_sheet.iter_rows(min_row=1, max_row=input_sheet.max_row, min_col=1, max_col=input_sheet.max_column):
    #     row_values = [cell.value for cell in row if cell.value is not None]
    #     print(row_values)
    #
    #     # 새로운 엑셀 파일 생성
    #     output_wb = openpyxl.Workbook()
    #     output_sheet = output_wb.active
    #
    #     if row_values is not None:
    #         output_sheet.append([row_values])
    #
    #         # 새로운 엑셀 파일 저장
    #         output_wb.save(output_excel_path)
    #         print('완료')

    # 선택한 열에서 한 줄씩 읽어와서 새로운 엑셀 파일에 쓰기
    # for row in input_sheet.iter_rows(min_row=1, max_row=input_sheet.max_row, min_col=column_index,
    #                                  max_col=column_index):
    for row in input_sheet.iter_rows(min_row=1, max_row=input_sheet.max_row, min_col=1, max_col=input_sheet.max_column):
        print('===============================')
        row_values = [cell.value for cell in row if cell.value is not None]
        print(row_values)
        row_values = [cell.value for cell in row if cell.value is not None]

        print(fr'row_values :{row_values}')
        print(fr'type(row_values) :{type(row_values)}')

        cell_value = row[0].value
        cell_value = row[0].value
        print(f"Cell Value row[0].value : {cell_value}")  # 이 부분을 추가하여 셀 값 출력
        print(fr'type(cell_value) :{type(cell_value)}')

        # 새로운 엑셀 파일 생성
        output_wb = openpyxl.Workbook()
        output_sheet = output_wb.active

        if cell_value is not None:
            output_sheet.append([cell_value])

    # 새로운 엑셀 파일 저장
    output_wb.save(output_excel_path)

    # # 새로운 엑셀 파일 생성
    # output_wb = openpyxl.Workbook()
    # output_sheet = output_wb.active
    #
    # # 선택한 열에서 한 줄씩 읽어와서 새로운 엑셀 파일에 쓰기
    # for row in input_sheet.iter_rows(min_row=1, max_row=input_sheet.max_row, min_col=column_index, max_col=column_index):
    #     cell_value = row[0].value
    #     print(f"Cell Value row[0].value : {cell_value}")  # 이 부분을 추가하여 셀 값 출력
    #
    #     if cell_value is not None:
    #         output_sheet.append([cell_value])
    #
    # # 새로운 엑셀 파일 저장
    # output_wb.save(output_excel_path)

if __name__ == "__main__":
    input_excel_path = fr"D:\data\vi\다국어문구_글씨 크게_test.xlsx"  # 읽어올 엑셀 파일 경로로 변경
    output_excel_path = fr"D:\text_file\라인명\vi\다국어문구_글씨 크게_test.xlsx"  # 저장할 새로운 엑셀 파일 경로로 변경
    column_index = 2  # 읽어올 열의 인덱스 (2는 B 열을 의미)

    read_and_write_excel(input_excel_path, output_excel_path, column_index)
