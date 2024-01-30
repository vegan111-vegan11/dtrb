import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import os
import time

def print_sheet_contents(sheet):
    # 각 셀의 내용 출력
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        row_values = [cell.value for cell in row if cell.value is not None]
        print(row_values)

def read_and_write_excel(input_excel_path, output_excel_path, column_index):
    # 기존 엑셀 파일 열기
    input_wb = openpyxl.load_workbook(input_excel_path)
    input_sheet = input_wb.active

    # 엑셀 파일의 내용 출력
    #print_sheet_contents(input_sheet)

    # for row in input_sheet.iter_rows(min_row=1, max_row=input_sheet.max_row, min_col=1, max_col=input_sheet.max_column):
    #     row_values = [cell.value for cell in row if cell.value is not None]
    #     print(row_values)
    #
    #     # 새로운 엑셀 파일 생성
    #     output_wb = openpyxl.Workbook()
    #     output_sheet = output_wb.active
    #
    #     if row_values is not None:
    #         output_sheet.append([row_values])
    #
    #         # 새로운 엑셀 파일 저장
    #         output_wb.save(output_excel_path)
    #         print('완료')

    # 선택한 열에서 한 줄씩 읽어와서 새로운 엑셀 파일에 쓰기
    # for row in input_sheet.iter_rows(min_row=1, max_row=input_sheet.max_row, min_col=column_index,
    #                                  max_col=column_index):
    for row_number, row in enumerate(
            input_sheet.iter_rows(min_row=1, max_row=input_sheet.max_row, min_col=1, max_col=input_sheet.max_column),
            1):
        print(f"Row {row_number}:")
        for cell_number, cell in enumerate(row, 1):
            print(f"  Cell {cell_number}: {cell.value}")



    #for row in input_sheet.iter_rows(min_row=1, max_row=input_sheet.max_row, min_col=1, max_col=input_sheet.max_column):
    for row_number, row in enumerate(input_sheet.iter_rows(min_row=1, max_row=input_sheet.max_row, min_col=1, max_col=input_sheet.max_column), 0):
        print('===============================')

        file_path = str(row_number).zfill(9)  # line_number를 9자리 숫자로 변환
        # 2초 동안 일시 정지
        time.sleep(0.02)
        # 파일명에 사용할 수 없는 문자 대체
        # file_name = replace_illegal_characters(file_name)

        # 파일 경로 생성
        output_excel_path = os.path.join(output_excel_path_folder, f'{file_path}.xlsx')


        # row_values = [cell.value for cell in row if cell.value is not None]
        # print(row_values)
        row_values = [cell.value for cell in row if cell.value is not None]

        print(fr'row_values :{row_values}')
        print(fr'type(row_values) :{type(row_values)}')

        #cell_value = row[0].value
        cell_value = row[0].value
        print(f"Cell Value row[0].value : {cell_value}")  # 이 부분을 추가하여 셀 값 출력
        print(fr'type(cell_value) :{type(cell_value)}')

        # 새로운 엑셀 파일 생성
        output_wb = openpyxl.Workbook()
        output_sheet = output_wb.active

        # if cell_value is not None:
        #     output_sheet.append([cell_value])
        if cell_value is not None:
            # 글씨 크기 72로 설정된 폰트 생성
            font = Font(size=72)
            font = Font(name='Noto Sans Thai Looped', size=72)
            # 새로운 셀에 값과 스타일 적용
            #new_cell = output_sheet.cell(row=len(output_sheet['A']) + 1, column=1, value=cell_value)
            #new_cell = output_sheet.cell(row=1, column=1, value=cell_value)
            # new_cell = output_sheet.cell(row=len(output_sheet['A']) + 0, column=1, value=cell_value)
            # new_cell.font = font
            max_length = 100
            max_length = 30
            max_length = 50
            max_length = 37
            # 셀에 대해 줄 바꿈 설정
            #new_cell.alignment = Alignment(wrap_text=True)
            #new_cell.alignment = Alignment(wrap_text=True)
            # 입력된 문자열을 일정 길이(max_length)마다 나누어 리스트로 변환
            parts = [cell_value[i:i + max_length] for i in range(0, len(cell_value), max_length)]
            print(f"parts : {parts}")  # 이 부분을 추가하여 셀 값 출력

            # 줄바꿈 문자('\n')을 이용하여 각 부분을 연결하여 새로운 문자열 생성
            #result = '\n'.join(parts)
            result = '\n'.join(parts)
            print(f"result : {result}")  # 이 부분을 추가하여 셀 값 출력
            cell_value = result
            print(f"cell_value 줄바꿈 후 : {cell_value}")  # 이 부분을 추가하여 셀 값 출력
            new_cell = output_sheet.cell(row=len(output_sheet['A']) + 0, column=1, value=cell_value)
            new_cell.font = font
            new_cell.alignment = Alignment(wrap_text=True)

            # 열의 너비 설정 (예: 20)
            output_sheet.column_dimensions['A'].width = 200
            output_sheet.column_dimensions['A'].width = 250
            output_sheet.column_dimensions['A'].width = 350
            output_sheet.column_dimensions['A'].width = 300

            output_sheet.row_dimensions[1].height = 350
            output_sheet.row_dimensions[1].height = 500  # 적절한 높이로 조절해보세요
            output_sheet.row_dimensions[1].height = 500  # 적절한 높이로 조절해보세요

            #output_sheet.row_dimensions[1].height = 30

            # 디렉터리 생성
            os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)

            # 새로운 엑셀 파일 저장
            output_wb.save(output_excel_path)

    # # 새로운 엑셀 파일 생성
    # output_wb = openpyxl.Workbook()
    # output_sheet = output_wb.active
    #
    # # 선택한 열에서 한 줄씩 읽어와서 새로운 엑셀 파일에 쓰기
    # for row in input_sheet.iter_rows(min_row=1, max_row=input_sheet.max_row, min_col=column_index, max_col=column_index):
    #     cell_value = row[0].value
    #     print(f"Cell Value row[0].value : {cell_value}")  # 이 부분을 추가하여 셀 값 출력
    #
    #     if cell_value is not None:
    #         output_sheet.append([cell_value])
    #
    # # 새로운 엑셀 파일 저장
    # output_wb.save(output_excel_path)

if __name__ == "__main__":
    input_excel_path = fr"D:\data\vi\다국어문구_글씨 크게_test.xlsx"  # 읽어올 엑셀 파일 경로로 변경
    output_excel_path = fr"D:\text_file\라인명\vi\다국어문구_글씨 크게_test.xlsx"  # 저장할 새로운 엑셀 파일 경로로 변경
    input_excel_path = fr"D:\data\vi\다국어문구_글씨 크게.xlsx"  # 읽어올 엑셀 파일 경로로 변경
    output_excel_path_folder = fr"D:\text_file\라인명\vi"  # 저장할 새로운 엑셀 파일 경로로 변경
    input_excel_path = fr"D:\data\vi\다국어문구_글씨 크게_test.xlsx"  # 읽어올 엑셀 파일 경로로 변경
    output_excel_path = fr"D:\text_file\라인명\vi\다국어문구_글씨 크게_test.xlsx"  # 저장할 새로운 엑셀 파일 경로로 변경
    column_index = 2  # 읽어올 열의 인덱스 (2는 B 열을 의미)

    read_and_write_excel(input_excel_path, output_excel_path_folder, column_index)
