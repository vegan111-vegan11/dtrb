# 라인명으로 이미지 저장
# 되는거
# 처음 이미지 저장
import os
import pyautogui
import time
import pygetwindow as gw

import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import subprocess

# lan = '태국어'
# font = 'NotoSansThaiLooped-Black'
# font = 'NotoSansThaiLooped-Bold'
# font = 'NotoSansThaiLooped-ExtraBold'
# font = 'NotoSansThaiLooped-ExtraLight'
# font = 'NotoSansThaiLooped-Light'
# font = 'NotoSansThaiLooped-Medium'
# font = 'NotoSansThaiLooped-Regular'
# font = 'NotoSansThaiLooped-SemiBold'
font = 'NotoSansThaiLooped-Thin'
# 글꼴 스타일 리스트 =
# [ 가늘게, 아주 가늘게, 보통, 중간, 약간 굵게, 굵게, 아주 굵게, 검정, 가는 기울임꼴, 아주 가는 기울임꼴, 기울임꼴, 중간 기울임꼴, 약간 굵은 기울임꼴, 굵은 기울임꼴, 아주 굵은 기울임꼴, 검은 기울임꼴]
# 글꼴 스타일 리스트 = [ 가늘게, 아주 가늘게, 보통, 중간, 약간 굵게, 굵게, 아주 굵게, 검정, 가는 기울임꼴, 아주 가는 기울임꼴, 기울임꼴, 중간 기울임꼴, 약간 굵은 기울임꼴, 굵은 기울임꼴, 아주 굵은 기울임꼴, 검은 기울임꼴]
# font_list = [ 가늘게, 아주 가늘게, 보통, 중간, 약간 굵게, 굵게, 아주 굵게, 검정, 가는 기울임꼴, 아주 가는 기울임꼴, 기울임꼴, 중간 기울임꼴, 약간 굵은 기울임꼴, 굵은 기울임꼴, 아주 굵은 기울임꼴, 검은 기울임꼴]
# for font in font_list:
#     print(f'font = \'{font}\'')
font_dict = {}
font_dict = {'검정': 'Black', '굵게': 'Bold', '아주 굵게': 'ExtraBold',
             '아주 가늘게': 'ExtraLight', '가늘게': 'Light', '중간': 'Medium',
             '보통': 'Regular', '약간 굵게': 'SemiBold', '가늘게': 'Thin'
             }

# font = '가늘게'
# font = '아주 가늘게'
# font = '보통'
# font = '중간'
# font = '약간 굵게'
# font = 'NotoSansThaiLooped-Black'
#font = '굵게'
#font = '아주 굵게'
#font = '검정'
# font = '가는 기울임꼴'
# font = '아주 가는 기울임꼴'
# font = '기울임꼴'
# font = '중간 기울임꼴'
# font = '약간 굵은 기울임꼴'
# font = '굵은 기울임꼴'
# font = '아주 굵은 기울임꼴'
# font = '검은 기울임꼴'

# 모든 열려있는 창 가져오기
windows = gw.getAllTitles()

excel_file_path = "D:\\text_file\\라인명\\test\\NotoSansThaiLooped-Thin\\000000028.xlsx"
os.system(f'start excel.exe {excel_file_path}')

x, y, width, height = 30, 60, 1800, 760  # 캡처하려는 화면 영역의 좌표와 크기
screenshot = pyautogui.screenshot(region=(x, y, width, height))

# lan = 'vi'
# output_folder = fr'D:\img\라인명\{lan}\{font}'
# line_number = 0
# output_file_path = os.path.join(output_folder, f'image-{line_number:09d}.jpg')


print('처음 염')
#time.sleep(1)

lan = '태국어'
lan = '베트남어'
lan = 'vi'
# 폴더 경로
# base_directory = r'C:\Users\TAMSystech\yjh\text_file\라인명\ttf'
# base_directory = r'C:\Users\TAMSystech\yjh\text_file\라인명2'

# 디렉토리를 반복하면서 디렉토리명 출력
# for directory_name in os.listdir(base_directory):
#     directory_path = os.path.join(base_directory, directory_name)


# print(f'directory_path : {directory_path}')
# #if(directory_name == 'LGBD'):
# # 디렉토리인 경우에만 출력
# if os.path.isdir(directory_path):

# 모든 창을 최소화
# for window in windows:
#     if window != "Program Manager":  # Windows의 Program Manager 창은 최소화하지 않도록 제외
#         app = gw.getWindowsWithTitle(window)
#         if app:
#             app[0].minimize()
print("모든 열려있는 창을 최소화했습니다.")
# 폴더 경로
# folder_path = fr'C:\Users\TAMSystech\yjh\text_file\라인명2\ttf\{directory_name}\{lan}'
# output_folder = fr'C:\Users\TAMSystech\yjh\img\라인명2\ttf\{directory_name}\{lan}'
# folder_path = fr'C:\Users\TAMSystech\yjh\text_file\라인명2\{lan}'
folder_path = fr'C:\Users\TAMSystech\yjh\text_file\라인명2\테스트\{lan}'
#folder_path = fr'C:\Users\TAMSystech\yjh\text_file\라인명\{lan}'
# 실전용
folder_path = fr'C:\Users\TAMSystech\yjh\text_file\라인명\{lan}\{font}'
folder_path = fr'D:\text_file\라인명\{lan}\{font}'
folder_path = fr'D:\text_file\라인명\test\{font}'
# 테스트용
#folder_path = fr'C:\Users\TAMSystech\yjh\text_file\라인명2\테스트\{lan}'
# 테스트용
#folder_path = fr'C:\Users\TAMSystech\yjh\text_file\라인명\{lan}\테스트'
#############################################################################

#folder_path = fr'C:\Users\TAMSystech\yjh\text_file\라인명2\테스트\{lan}'
#folder_path = fr'C:\Users\TAMSystech\yjh\text_file\라인명\{lan}\{font}'
#folder_path = fr'C:\Users\TAMSystech\yjh\text_file\라인명\{lan}\{font}'

output_folder = fr'C:\Users\TAMSystech\yjh\img\라인명\{lan}\{font}'
# output_folder = fr'C:\Users\TAMSystech\yjh\img\라인명\{lan}\테스트'
output_folder = fr'C:\Users\TAMSystech\yjh\img\라인명\{lan}\{font}'
output_folder = fr'D:\img\라인명\{lan}\{font}'
#output_folder = fr'C:\Users\TAMSystech\yjh\img\라인명\{lan}\테스트'
# output_folder = fr'C:\Users\TAMSystech\yjh\img\라인명2\테스트\{lan}\{font}'
# folder_path = r'C:\Users\TAMSystech\yjh\text_file\테스트\테스트'
# output_folder = fr'C:\Users\TAMSystech\yjh\img\ttf2\{directory_name}\{lan}'

# 디렉토리가 없으면 생성
if not os.path.exists(output_folder):
    os.makedirs(output_folder)
pyautogui.FAILSAFE = False
# 폴더 내의 모든 txt 파일 가져오기
txt_files = [f for f in os.listdir(folder_path) if f.endswith('.txt')]
txt_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
# 폴더 내의 모든 txt 파일 가져오기
# txt_files = [f for f in os.listdir(folder_path) if f.endswith('.txt') and any(str(num) in f for num in [202, 203, 204, 211, 220])]
print(f'!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!txt_files : {txt_files}')
line_number = 0
# 텍스트 파일을 열어서 화면 캡처 및 이미지 저장
# for txt_file in txt_files:
#     txt_file_path = os.path.join(folder_path, txt_file)

input_excel_path = fr"D:\text_file\라인명\test\NotoSansThaiLooped-Thin\000000028.xlsx"  # 읽어올 엑셀 파일 경로로 변경

input_wb = openpyxl.load_workbook(input_excel_path)
input_sheet = input_wb.active

print('열려야 함======================')
excel_ficle_path = "D:\\text_file\\라인명\\test\\NotoSansThaiLooped-Thin\\000000028.xlsx"

#os.system(f'start excel.exe {excel_ficle_path}')


#time.sleep(0.1)

for row_number, row in enumerate(
        input_sheet.iter_rows(min_row=1, max_row=input_sheet.max_row, min_col=1, max_col=input_sheet.max_column), 0):
    print('===============================')

    print(f"Row {row_number}:")
    for cell_number, cell in enumerate(row, 1):
        print(f"  Cell {cell_number}: {cell.value}")

    file_path = str(row_number).zfill(9)  # line_number를 9자리 숫자로 변환
    # 2초 동안 일시 정지
    #time.sleep(0.02)
    print(f'!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!file_path : {file_path}')

    # 모든 열려있는 창 가져오기
    #windows = gw.getAllTitles()
    print(f'!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!1 : ')
    # 모든 창을 최소화
    #     for window in windows:
    #         #print(f'!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!window : {window}')
    #         if window != "Program Manager":  # Windows의 Program Manager 창은 최소화하지 않도록 제외
    #             #print(f'!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!3 : ')
    #             app = gw.getWindowsWithTitle(window)
    #             print(f'!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!app : {app}')
    #             #print(f'!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!4 : ')
    #             if app:
    #                 #print(f'!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!5 : ')
    #                 app[0].minimize()
    #                 #print(f'!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!6 : ')

    # 모든 창을 최소화
    # for window in windows:
    #     if window != "Program Manager":  # Windows의 Program Manager 창은 최소화하지 않도록 제외
    #         app = gw.getWindowsWithTitle(window)
    #         if app:
    #             app[0].minimize()
    print("두번째 모든 열려있는 창을 최소화했습니다.")

    # txt 파일을 메모장으로 엽니다 (Windows 기준)
    # os.system(f'start notepad.exe {txt_file_path}')
    # 노트패드로 파일 열기
    # subprocess.run(['notepad.exe', txt_file_path])
    # 노트패드로 파일 열기
    # subprocess.run(['notepad.exe', txt_file_path])
    # subprocess.run(['notepad.exe', txt_file_path])
    # notepadpp_path = r'C:\Program Files\Notepad++\notepad++.exe'
    print(f'!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!7 : ')
    print(f'!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!8 : ')
    #os.system(f'start notepad.exe {txt_file_path}')
    # input_wb = openpyxl.load_workbook(txt_file_path)
    # input_sheet = input_wb.active
    # Notepad++로 파일 열기
    # subprocess.run([notepadpp_path, txt_file_path])
    print(f'!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!9 : ')
    # C:\Program Files\Notepad++\notepad++.exe
    #time.sleep(0.1)
    #time.sleep(1)
    #print(f'노트패드로 파일 열기txt_file_path : {txt_file_path}')

    # 메모장 창이 열릴 때까지 대기합니다 (시간을 필요에 따라 조정하세요)
    # #time.sleep(1)
    # 메모장 창 최대화 (Windows 창 최대화 단축키)
    pyautogui.hotkey('win', 'up')  # 창을 최대화합니다
    print('창 최대하 함===============')
    #time.sleep(0.1)
    #time.sleep(1)
    # 글자 크기 조정 (키보드 조작을 통해 Ctrl + "-" 키를 누름)
    # pyautogui.hotkey('ctrl', '-')
    # time.sleep(0.5)  # 조절에 시간이 필요한 경우 대기시간 추가
    # 커서 숨기기
    # pyautogui.moveTo(-1, -1)  # 커서를 (1, 1)로 이동하여 숨깁니다
    # time.sleep(0.5)  # 조절에 시간이 필요한 경우 대기시간 추가
    # 화면 캡처를 캡처합니다
    # x, y, width, height = 10, 48, 1800, 120  # 캡처하려는 화면 영역의 좌표와 크기
    # x, y, width, height = 6, 48, 1800, 80  # 캡처하려는 화면 영역의 좌표와 크기
    x, y, width, height = 19, 242, 1800, 760  # 캡처하려는 화면 영역의 좌표와 크기
    x, y, width, height = 19, 242, 100, 100  # 캡처하려는 화면 영역의 좌표와 크기
    x, y, width, height = 19, 242, 1000, 760  # 캡처하려는 화면 영역의 좌표와 크기
    x, y, width, height = 1, 1, 1000, 760  # 캡처하려는 화면 영역의 좌표와 크기
    x, y, width, height = 10, 1, 1000, 760  # 캡처하려는 화면 영역의 좌표와 크기
    x, y, width, height = 30, 60, 1000, 760  # 캡처하려는 화면 영역의 좌표와 크기
    x, y, width, height = 30, 60, 1800, 760  # 캡처하려는 화면 영역의 좌표와 크기
    screenshot = pyautogui.screenshot(region=(x, y, width, height))
    #time.sleep(1)
    # 이미지를 저장합니다
    # time.sleep(0.5)  # 조절에 시간이 필요한 경우 대기시간 추가

    output_file_path = os.path.join(output_folder, f'image-{line_number:09d}.jpg')

    # 이미 파일이 존재하는지 체크
    if os.path.exists(output_file_path):
        # print(f'파일이 이미 존재합니다: {output_file_path}')
        os.remove(output_file_path)

        # print(f'기존 파일 삭제: {output_file_path}')

        screenshot.save(output_file_path)
        print(f'스크린샷 이미지를 저장했습니다: {output_file_path}')
        # time.sleep(0.5)  # 조절에 시간이 필요한 경우 대기시간 추가
        # print(f'{line_number} 개의 txt 파일의 스크린샷 이미지를 저장했습니다.')
        # 여기에 필요한 처리를 추가하세요 (덮어쓰기, 새로운 이름 사용 등)
    else:
        screenshot.save(output_file_path)
        print(f'스크린샷 이미지를 저장했습니다2: {output_file_path}')

    # for line in open(txt_file_path):
    # for line in open(txt_file_path, 'r', encoding='utf-8'):
    #     # line_number += 1
    #     # print(f'line_number : {line_number}')
    #     # print(f'txt_file : {txt_file}')
    #     # txt_file = txt_file[:-4]
    #     # txt_file = int(txt_file)
    #     # output_file_path = os.path.join(output_folder, f'image-{txt_file:09d}.jpg')
    #     line_number = int(txt_file[:-4])
    #     print(f'!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!line_number : {line_number}')
    #
    #     output_file_path = os.path.join(output_folder, f'image-{line_number:09d}.jpg')
    #     print(f'!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!output_file_path : {output_file_path}')
    #     # screenshot.save(output_file_path)
    #     # print(f'스크린샷 이미지를 저장했습니다: {output_file_path}')
    #
    #     # 이미 파일이 존재하는지 체크
    #     if os.path.exists(output_file_path):
    #         # print(f'파일이 이미 존재합니다: {output_file_path}')
    #         os.remove(output_file_path)
    #
    #         # print(f'기존 파일 삭제: {output_file_path}')
    #
    #         screenshot.save(output_file_path)
    #         print(f'스크린샷 이미지를 저장했습니다: {output_file_path}')
    #         # time.sleep(0.5)  # 조절에 시간이 필요한 경우 대기시간 추가
    #         # print(f'{line_number} 개의 txt 파일의 스크린샷 이미지를 저장했습니다.')
    #         # 여기에 필요한 처리를 추가하세요 (덮어쓰기, 새로운 이름 사용 등)
    #     else:
    #         screenshot.save(output_file_path)
    #         print(f'스크린샷 이미지를 저장했습니다2: {output_file_path}')

    # 메모장 창을 닫습니다
    #os.system('taskkill /f /im notepad.exe')
    # line_number += 1
print(f'{len(txt_files)} 개의 txt 파일의 스크린샷 이미지를 저장했습니다.')