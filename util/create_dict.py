import os
import openpyxl

# Excel 파일 경로
excel_file_path = r'D:\data\다국어문구 - 복사본.xlsx'

# 결과를 저장할 파일 경로
output_file_path = r'C:\Users\TAMSystech\yjh\ipynb\TextRecognitionDataGenerator\trdg\dicts\전체언어텍스트파일\th_1212.txt'

# 찾을 열 이름
thai_column_name = "태국어"  # 열의 실제 이름에 따라 변경

# Excel 파일 열기
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook.active

# 열 이름 찾기
thai_column_index = None
for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):  # 두 번째 행만 확인 (max_row=2)
    print(f' row : { row}')
    for col_idx, header in enumerate(row, start=1):
        print(f' idx : { idx}')
        print(f' header : { header}')
        if header == thai_column_name:
            thai_column_index = col_idx
            print(f' thai_column_index : { thai_column_index}')
            break

if thai_column_index is not None:
    # 파일 열기 및 쓰기
    with open(output_file_path, 'w', encoding='utf-8') as output_file:
        # 각 행에 대해 반복
        for row in sheet.iter_rows(min_row=2, values_only=True):  # 세 번째 행부터 데이터가 시작됨 (min_row=3)
            print(f' thai_text 건너뛰기 전 row : { row}')
            thai_text = str(row[thai_column_index - 1]).strip()  # 태국어 열 값 가져오기 및 공백 제거
            print(f' thai_text 건너뛰기 전 thai_text : { thai_text}')
            # 빈 문자열이나 공백인 경우 건너뛰기
            if thai_text:
                print(f' thai_text : { thai_text}')
                output_file.write(thai_text + '\n')  # 파일에 쓰기

    print(f"태국어 텍스트가 {output_file_path}에 성공적으로 저장되었습니다.")
else:
    print(f"태국어 열을 찾을 수 없습니다. 열 이름을 확인하세요.")


input_file_path = r'C:\Users\TAMSystech\yjh\ipynb\TextRecognitionDataGenerator\trdg\dicts\전체언어텍스트파일\th_1212_2.txt'
output_file_path = r'C:\Users\TAMSystech\yjh\ipynb\TextRecognitionDataGenerator\trdg\dicts\전체언어텍스트파일\th_1212_3.txt'

# 파일 읽기
with open(input_file_path, 'r', encoding='utf-8') as input_file:
    lines = input_file.readlines()

# 빈 문자열 제거
non_empty_lines = [line.strip() for line in lines if line.strip()]

# 결과를 새로운 파일에 저장
with open(output_file_path, 'w', encoding='utf-8') as output_file:
    output_file.write('\n'.join(non_empty_lines))

print(f"빈 문자열이 제거된 텍스트가 {output_file_path}에 저장되었습니다.")
