import openpyxl
import pyautogui
import os
import time

# 엑셀 파일 경로
excel_file_path = r'D:\data\다국어문구_글씨 크게.xlsx'

# 스크린샷 저장 폴더
output_folder = r'C:\Users\TAMSystech\yjh\img\라인명2\베트남어'
os.makedirs(output_folder, exist_ok=True)

# 엑셀 파일 열기
wb = openpyxl.load_workbook(excel_file_path)
sheet = wb.active

# 한 줄의 높이 가져오기
row_height = sheet.row_dimensions[1].height
row_height = 100

# 스크롤 간격 설정
scroll_distance = int(row_height)  # 한 줄의 높이만큼 스크롤
scroll_distance = int(row_height)  # 한 줄의 높이만큼 스크롤

# 세로는 한 줄의 높이, 가로는 1000으로 설정
region_width = 1000
region_height = int(row_height)

# 시작 위치 설정
start_x, start_y = 10, 10

# 엑셀 파일 열기
wb = openpyxl.load_workbook(excel_file_path)
sheet = wb.active

# 엑셀 파일 열기
os.startfile(excel_file_path)


# 행의 개수만큼 반복
#for row_number in range(1, sheet.max_row + 1):
for row_number in range(1, 1 + 1):

    # 스크롤
    pyautogui.scroll(scroll_distance)
    pyautogui.scroll(200)

    # 대기 시간
    time.sleep(1)

    # 스크롤 후 영역 설정하여 스크린샷 찍기
    screenshot = pyautogui.screenshot(region=(start_x, start_y, region_width, region_height))

    # 스크린샷 파일 저장
    screenshot_path = os.path.join(output_folder, f'screenshot_{row_number}.jpg')
    screenshot.save(screenshot_path)

# 엑셀 파일 닫기
wb.close()
