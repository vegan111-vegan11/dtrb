import openpyxl
import pyautogui
import os
import time

# 엑셀 파일 경로
excel_file_path = r'D:\data\다국어문구_글씨 크게.xlsx'

# 스크린샷 저장 폴더
output_folder = r'C:\Users\TAMSystech\yjh\img\라인명2\베트남어'
os.makedirs(output_folder, exist_ok=True)

# 한 번에 스크롤하는 거리 (조절 가능)
scroll_distance = 300
scroll_distance = 600
scroll_distance = 100
scroll_distance = 200
scroll_distance = 50
scroll_distance = 100
scroll_distance = 150
scroll_distance = 100
scroll_distance = 150

# 대기 시간 설정 (스크롤이 발생하면 기다려야 함)
scroll_wait_time = 1

# 엑셀 파일 열기
wb = openpyxl.load_workbook(excel_file_path)
sheet = wb.active

# 엑셀 파일 열기
os.startfile(excel_file_path)

# 한 줄의 높이 가져오기
row_height = sheet.row_dimensions[1].height
# 첫 번째 행의 높이 가져오기
first_row_number = 1
row_height = sheet.row_dimensions[first_row_number].height
# 특정 행 번호 설정 (예: 1)
target_row_number = 1

# 특정 행에 데이터가 있는 경우에만 높이 출력
if sheet.cell(row=target_row_number, column=1).value:
    row_height = sheet.row_dimensions[target_row_number].height
    print(f"한 줄의 높이: {row_height}")
else:
    print(f"특정 행({target_row_number})에 데이터가 없습니다.")


# 엑셀 파일 닫기
#wb.close()

# 대기 시간을 고려하여 스크롤 및 스크린샷 찍기
#for row_number in range(1, sheet.max_row + 1):
for row_number in range(1, 1 + 1):
    # 스크롤
    pyautogui.scroll(scroll_distance)

    # 대기 시간
    time.sleep(scroll_wait_time)

    # 스크린샷 찍기
    screenshot = pyautogui.screenshot()
    screenshot = pyautogui.screenshot(region=(100, 200, 800, 600))
    screenshot = pyautogui.screenshot(region=(100, 200, 800, 600))
    screenshot = pyautogui.screenshot(region=(10, 20, 800, 200))
    screenshot = pyautogui.screenshot(region=(10, 20, 800, 50))

    screenshot_path = os.path.join(output_folder, f'screenshot_{row_number}.jpg')
    screenshot.save(screenshot_path)

# 프로그램 종료 전에 엑셀 파일 닫기 (추가된 부분)
#wb.close()

# 여러 번 스크롤 아래로
for row_number in range(3):  # 예시로 5번 스크롤
    pyautogui.scroll(-scroll_distance)
    time.sleep(scroll_wait_time)

    # 스크린샷 찍기
    screenshot = pyautogui.screenshot()
    screenshot = pyautogui.screenshot(region=(100, 200, 800, 600))
    screenshot = pyautogui.screenshot(region=(100, 200, 1800, 600))

    screenshot_path = os.path.join(output_folder, f'screenshot_{row_number}.jpg')
    screenshot.save(screenshot_path)
